#4nm19is120 4nm19is106 4nm19is148
#This program was developed on 15-04-2022
#openpyxl and xlrd should be installed


from asyncio.windows_events import NULL
from turtle import goto
import xlrd
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from pathlib import Path


toClose = False
i=1



print("\t\tThis is Library Attendence Management System")
print("\t\tThis System has been Developed by Information Science Department: \n\t\tDeveloped By Pratheek Shankar, Akshay Padmashali and Srinivas Shanti")
print("\t\tNOTE: A folder named Entries should be present in the current folder")
print("\t\tType cancel to exit (all small)\n\n\n")


wb1 = xlrd.open_workbook('data.xls')
sheet = wb1.sheet_by_index(0)

today = datetime.now().strftime("%m_%Y")

file_name = os.getcwd()
file_name = file_name+"/Entries/"+today+".xlsx"

path = Path(file_name)

if path.is_file():
    wb = openpyxl.load_workbook(file_name)
    sheet1 = wb.active
    
    i = sheet1.max_row
    
else:
    print('Creating a new file')
    
    wb = Workbook()
    sheet1 = wb.active
    

    sheet1.cell(row= i , column = 1).value = "USN"    
    sheet1.cell(row= i , column = 2).value = "Name"
    sheet1.cell(row= i , column = 3).value = "Department"
    sheet1.cell(row= i , column = 4).value = "Sem"
    sheet1.cell(row= i , column = 5).value = "In-Time"
    sheet1.cell(row= i , column = 6).value = "Date"



while (toClose == False):
    
    print("\n\nEnter your USN: ")
    
    barReader = input()
    
    outTime = False
    
    isValid = False

    if barReader == "cancel":
        break
    
    if barReader.isnumeric():
        barReader = int(barReader)
        
    elif barReader.islower():
        barReader = barReader.upper()

    now = datetime.now()
    
    for row_num in range(sheet.nrows):
        row_value = sheet.row_values(row_num)
        if row_value[0] == barReader:
    
            i+=1
            
            isValid = True
            
            today = datetime.now().strftime("%d_%m_%Y")
            
            current_time = now.strftime("%H:%M:%S")
            
    #         for row1_num in range(sheet1.max_row):
    #             if sheet1.cell(row = row1_num + 1 , column = 1).value == barReader:
    #                 if sheet1.cell(row = row1_num + 1 , column = 6).value == NULL:
    #                     if sheet1.cell(row1_num + 1 , column = 7).value == today:
    #                         outTime = True
    #                         sheet1.cell(row= row1_num + 1 , column = 6).value = current_time
    #                         wb.save(file_name)
    #                         i-=1
    #                         print("Data has been recorded to outTime\n USN: ",barReader,"Name: ",row_value[1])
    #                         break
                        
                    

    #         if outTime == True:
    #             continue
            
            
                
            # current_time = now.strftime("%H:%M:%S")
            # today = datetime.now().strftime("%d_%m_%Y")
                
            sheet1.cell(row= i , column = 1).value = barReader
            sheet1.cell(row= i , column = 2).value = row_value[1]   
            sheet1.cell(row= i , column = 3).value = row_value[2]
            sheet1.cell(row= i , column = 4).value = row_value[3]
            sheet1.cell(row= i , column = 5).value = current_time
            sheet1.cell(row= i , column = 6).value = today

            wb.save(file_name)
                
            print("Data has been recorded\n\t\tUSN: ",barReader,"\n\t\tName: ",row_value[1],"")
            break
        
    if isValid == False:
        print("\t!!!!!!!!!!!!!!!Error!!!!!!!!!!!!!!! \nPlease Try Again")